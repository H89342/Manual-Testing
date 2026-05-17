#!/usr/bin/env python3
"""
Convert the v4 CSV deliverables into a formatted .xlsx workbook.

Usage:
    cd "Telemax_Battery Prediction feature"
    python3 -m pip install --user openpyxl   # one-time
    python3 build_xlsx_from_csv.py

Reads:  testcases_battery_health_v4.csv
        testcases_battery_health_v4_CQ.csv
Writes: testcases_battery_health_v4.xlsx
"""
import csv
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
TC_CSV  = os.path.join(HERE, "testcases_battery_health_v4.csv")
CQ_CSV  = os.path.join(HERE, "testcases_battery_health_v4_CQ.csv")
XLSX    = os.path.join(HERE, "testcases_battery_health_v4.xlsx")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required. Install with:  python3 -m pip install --user openpyxl")


HEADER_FILL  = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
GROUP_FILL   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
THIN         = Side(border_style="thin", color="B4B4B4")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TC_COL_WIDTHS = [12, 38, 50, 60, 40, 60, 60, 32, 40, 12, 10, 14, 28, 30]
TC_TOP_COLS   = {3, 5, 6, 7}  # C, E, F, G — multi-line cells get top alignment

CQ_COL_WIDTHS = [10, 18, 40, 60, 60, 10, 14, 14, 40, 12, 14]


def is_pending(notes_cell: str) -> bool:
    return "PENDING" in (notes_cell or "")


def write_tcs(wb):
    ws = wb.active
    ws.title = "Test Cases"
    with open(TC_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return
    headers = rows[0]
    data = rows[1:]

    # header row
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, w in enumerate(TC_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        pending = is_pending(row[13] if len(row) > 13 else "")
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci in TC_TOP_COLS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if pending:
                cell.fill = ORANGE_FILL
        # estimate row height
        max_lines = max(
            (row[5].count("\n") + 1) if len(row) > 5 else 1,
            (row[6].count("\n") + 1) if len(row) > 6 else 1,
            (row[4].count("\n") + 1) if len(row) > 4 else 1,
            (row[8].count("\n") + 1) if len(row) > 8 else 1,
            3,
        )
        ws.row_dimensions[ri].height = max(60, min(420, 18 * max_lines + 8))
    print(f"[xlsx] Test Cases sheet: {len(data)} rows")


def write_cqs(wb):
    ws = wb.create_sheet("Clarify Requirements")
    with open(CQ_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return
    headers = rows[0]
    data = rows[1:]

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, w in enumerate(CQ_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci in (3, 4, 5):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = ORANGE_FILL
        ws.row_dimensions[ri].height = 90
    print(f"[xlsx] Clarify Requirements sheet: {len(data)} rows")


def write_summary(wb):
    """Summary sheet derived from the Test Cases CSV."""
    ws = wb.create_sheet("Summary")
    sum_headers = ["TC ID", "Screen / Section", "Title", "Type", "Priority", "Status", "Notes"]
    widths      = [12, 50, 60, 14, 10, 12, 30]
    for ci, h in enumerate(sum_headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    with open(TC_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    headers = rows[0]
    data    = rows[1:]

    col_idx = {h: i for i, h in enumerate(headers)}
    take    = [
        col_idx["TC ID"],
        col_idx["Screen / Section"],
        col_idx["Title"],
        col_idx["Type"],
        col_idx["Priority"],
        col_idx["Status"],
        col_idx["Notes"],
    ]
    for ri, row in enumerate(data, start=2):
        pending = is_pending(row[col_idx["Notes"]])
        for ci, src in enumerate(take, start=1):
            cell = ws.cell(row=ri, column=ci, value=row[src])
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if pending:
                cell.fill = ORANGE_FILL
        ws.row_dimensions[ri].height = 36


def main():
    wb = openpyxl.Workbook()
    write_tcs(wb)
    write_cqs(wb)
    write_summary(wb)
    wb.save(XLSX)
    print(f"[done] wrote {XLSX}")


if __name__ == "__main__":
    main()
