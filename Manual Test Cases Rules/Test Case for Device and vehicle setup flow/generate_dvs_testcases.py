import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────
NAVY       = "1F4E79"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "DEEAF1"
WHITE      = "FFFFFF"
ORANGE     = "FFC000"
RED_PHASE  = "C00000"
GREEN_OK   = "375623"
GREY_TEXT  = "595959"

# ── Helpers ─────────────────────────────────────────────────────
def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=10, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def phase_sep(ws, row, label, col_count=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font    = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill    = fill(RED_PHASE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border  = thin_border()
    ws.row_dimensions[row].height = 18

def write_tc(ws, row, tc):
    cols = ["tc_id","title","preconditions","steps","expected",
            "test_data","postconditions","status","priority","type","environment","notes"]
    for c, key in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=tc.get(key, ""))
        cell.border    = thin_border()
        cell.alignment = wrap_align()
        is_orange = tc.get("pending", False)
        if is_orange:
            cell.fill = fill(ORANGE)
            cell.font = body_font(bold=(c in (1, 2)))
        elif c == 1:
            cell.font = body_font(bold=True, color=NAVY)
        elif c in (8,):   # Status
            cell.font = body_font(bold=True)
        else:
            cell.font = body_font()
    ws.row_dimensions[row].height = 90

# ════════════════════════════════════════════════════════════════
#  TEST CASE DATA
# ════════════════════════════════════════════════════════════════
NL = "\n"

TESTCASES = [

# ── PHASE 1: Device Connection ───────────────────────────────────
{
"tc_id": "DVS-001",
"title": "Verify displaying Welcome to Vehicle Onboarding screen with all five phases listed",
"preconditions": (
    "- User is logged into the Telemax mobile app\n"
    "- User has not started Vehicle Setup before\n"
    "- Device: iOS 16+ or Android 12+\n"
    "- App version: latest staging build"
),
"steps": (
    "1. Open Telemax app and tap 'Add Vehicle' or launch Vehicle Setup\n"
    "2. Observe the Welcome screen content\n"
    "3. Observe the list of phases displayed\n"
    "4. Tap 'Cancel'"
),
"expected": (
    "1. Welcome screen loads with title 'Welcome to Telemax Vehicle Onboarding'\n"
    "2. Subtitle: 'Complete these 5 steps to get your vehicle connected and ready to drive'\n"
    "3. All 5 phases listed: Phase 1 Device Connection, Phase 2 Rapid Photo Capture, Phase 3 Start Engine & Sync, Phase 4 Vehicle Review, Phase 5 Health Check\n"
    "4. 'Get Started' and 'Cancel' buttons are visible\n"
    "4. Tapping Cancel exits setup without saving"
),
"test_data": "app_account: testuser@example.com\napp_env: Staging",
"postconditions": "- No setup record is created when user cancels from Welcome screen",
"status": "Not Run", "priority": "High", "type": "UI",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-002",
"title": "Verify completing connectivity check successfully when mobile device has active internet connection",
"preconditions": (
    "- User is on the Welcome screen\n"
    "- Mobile device has an active LTE / Wi-Fi connection\n"
    "- Device: iOS 16+ or Android 12+"
),
"steps": (
    "1. Tap 'Get Started' on the Welcome screen\n"
    "2. Observe the Connectivity Check screen\n"
    "3. Observe the connection status\n"
    "4. Tap 'Start Setup'"
),
"expected": (
    "1. Connectivity Check screen loads with title 'Connectivity Check'\n"
    "2. Status shows 'Status: Connected' with green indicator\n"
    "3. Signal strength and network info are displayed\n"
    "4. 'Start Setup' and 'Exit Setup' buttons are visible\n"
    "4. Tapping 'Start Setup' advances to Phase 1 QR scan flow"
),
"test_data": "network: LTE or Wi-Fi active\napp_account: testuser@example.com",
"postconditions": "- User advances to 'How to Scan QR Code' screen",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-003",
"title": "Verify displaying connection required error when mobile device has no internet connection during connectivity check",
"preconditions": (
    "- User is on the Welcome screen\n"
    "- Mobile device has NO active internet connection (airplane mode or no signal)\n"
    "- Device: iOS 16+ or Android 12+"
),
"steps": (
    "1. Disable mobile data and Wi-Fi on the device\n"
    "2. Tap 'Get Started' on the Welcome screen\n"
    "3. Observe the Connectivity Check screen\n"
    "4. Tap 'Wait For Signal'\n"
    "5. Tap 'Continue Anyway'"
),
"expected": (
    "1. Connectivity Check screen shows 'Connection Required' error state with red indicator\n"
    "2. Message: 'Verifying your mobile device has a stable internet connection for the vehicle onboarding process'\n"
    "3. Error banner: 'Connection Required. Your device requires an internet connection for photo processing and vehicle look-ups. Please check your signal before continuing.'\n"
    "4. 'Wait For Signal' keeps user on the screen waiting\n"
    "5. 'Continue Anyway' allows user to proceed past connectivity check despite no connection\n"
    "5. 'Exit Setup' button is visible"
),
"test_data": "network: No connection (airplane mode)\napp_account: testuser@example.com",
"postconditions": "- User stays on Connectivity Check or proceeds with degraded mode if 'Continue Anyway' is tapped",
"status": "PENDING", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Blocked by CQ-001 — exact behaviour when 'Continue Anyway' is tapped without internet is not defined in Figma",
"pending": True
},

{
"tc_id": "DVS-004",
"title": "Verify displaying How to Scan QR Code instruction screen before opening the camera scanner",
"preconditions": (
    "- User has passed Connectivity Check\n"
    "- Device camera permission is granted\n"
    "- Device: iOS 16+ or Android 12+"
),
"steps": (
    "1. Complete Connectivity Check and tap 'Start Setup'\n"
    "2. Observe the 'How to Scan QR Code' screen\n"
    "3. Observe all content on screen\n"
    "4. Tap 'Watch Installation Video'\n"
    "5. Return and tap 'CAN\\'T SCAN? ENTER IMEI MANUALLY'\n"
    "6. Return and tap 'Start Scan'"
),
"expected": (
    "1. Screen loads with title 'How to Scan QR Code' and label 'PHASE 01 / 05'\n"
    "2. Instructions: 'Your tracker has two QR codes. Scan the one next to the IMEI number. It is a 15 digit number starting with 86512'\n"
    "3. Device image shows QR code location highlighted with red circle\n"
    "4. 'Watch Installation Video' link is tappable and opens video\n"
    "5. 'CAN\\'T SCAN? ENTER IMEI MANUALLY' link navigates to manual IMEI entry\n"
    "6. 'Start Scan' opens camera scanner on Scan QR Code screen\n"
    "6. 'Save & Exit' button is visible"
),
"test_data": "app_account: testuser@example.com",
"postconditions": "- Camera scanner opens when 'Start Scan' is tapped",
"status": "Not Run", "priority": "Medium", "type": "UI",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": "Updated screen per design update 24/04/2026 — How to Scan screen now shows before the actual scanner"
},

{
"tc_id": "DVS-005",
"title": "Verify scanning QR Code successfully with the correct IMEI QR code on the tracker device",
"preconditions": (
    "- User is on the Scan QR Code camera screen\n"
    "- Physical Teltonika tracker device with valid QR code is available\n"
    "- The tracker IMEI is registered to the test account\n"
    "- Camera permission is granted\n"
    "- Good lighting conditions"
),
"steps": (
    "1. Position camera over the QR code next to the IMEI number on the tracker\n"
    "2. Hold steady within the scanning frame\n"
    "3. Observe result after successful scan"
),
"expected": (
    "1. Camera frame is active and scanning\n"
    "2. QR code is detected within the frame\n"
    "3. App automatically reads the IMEI from the QR code\n"
    "3. User advances to 'Connect Tracker' screen showing tracker details and GPS optimisation info"
),
"test_data": (
    "tracker_imei: 860211234567890\n"
    "qr_code: QR code on tracker device next to IMEI label\n"
    "account: testuser@example.com"
),
"postconditions": (
    "- IMEI is captured and validated against the account\n"
    "- User is on the Connect Tracker confirmation screen"
),
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-006",
"title": "Verify remaining on Scan QR Code screen when a QR code of incorrect format (serial number) is scanned",
"preconditions": (
    "- User is on the Scan QR Code camera screen\n"
    "- Tracker device has two QR codes: IMEI QR and serial number QR\n"
    "- Camera permission is granted"
),
"steps": (
    "1. Intentionally scan the wrong QR code (serial number format, not IMEI format)\n"
    "2. Observe the on-screen response\n"
    "3. Re-position camera to scan correct QR code"
),
"expected": (
    "1. Red error banner appears on the Scan QR Code screen: 'Incorrect QR Code Scanned — Please scan the QR code next to the IMEI number. It is a 15 digit number starting with 86512'\n"
    "2. User remains on the Scan QR Code screen (does NOT navigate to Device Not Found)\n"
    "3. Camera scanner remains active for retry\n"
    "3. 'CAN\\'T SCAN? ENTER IMEI MANUALLY' and 'Exit Setup' links remain visible"
),
"test_data": (
    "wrong_qr: Serial number QR code on tracker (different format from IMEI QR)\n"
    "expected_format: 15-digit number starting with 86512"
),
"postconditions": "- User stays on Scan QR Code screen; no navigation to Device Not Found",
"status": "Not Run", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Design update 24/04/2026: incorrect QR (serial number format) now keeps user on scan page instead of going to Device Not Found"
},

{
"tc_id": "DVS-007",
"title": "Verify displaying Device Not Found screen when scanned IMEI is not registered to the account",
"preconditions": (
    "- User is on the Scan QR Code camera screen\n"
    "- Tracker IMEI exists physically but is NOT registered to the test account\n"
    "- Camera permission is granted"
),
"steps": (
    "1. Scan a valid IMEI-format QR code for a tracker not registered to the account\n"
    "2. Observe the result screen\n"
    "3. Tap 'Try Again'\n"
    "4. Tap 'Contact Support' on the second attempt"
),
"expected": (
    "1. 'Device Not Found' screen appears\n"
    "2. Message: 'The IMEI provided is not registered to your account. Please check the code or contact support.'\n"
    "3. Buttons visible: 'Try Again', 'Contact Support', 'Exit Setup'\n"
    "3. Tapping 'Try Again' returns user to QR scan screen\n"
    "4. Tapping 'Contact Support' shows 'Successfully notified support' confirmation banner"
),
"test_data": (
    "tracker_imei: 860219999999999\n"
    "status: IMEI valid format but not registered to testuser@example.com"
),
"postconditions": "- Support notification is sent when 'Contact Support' is tapped",
"status": "Not Run", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-008",
"title": "Verify entering IMEI manually and connecting tracker successfully with a valid 15-digit IMEI",
"preconditions": (
    "- User is on the 'How to Scan QR Code' screen or Scan QR Code screen\n"
    "- Tracker IMEI is registered to the test account\n"
    "- User has the IMEI number available (from label on device)"
),
"steps": (
    "1. Tap 'CAN\\'T SCAN? ENTER IMEI MANUALLY'\n"
    "2. Observe the 'Enter IMEI' screen\n"
    "3. Tap the TRACKER IDENTIFICATION NUMBER field\n"
    "4. Enter a valid 15-digit IMEI starting with 86512\n"
    "5. Tap 'Continue'"
),
"expected": (
    "1. 'Enter IMEI' screen loads with title and field counter showing x/15\n"
    "2. Instruction: 'You can find the IMEI on the front of your tracker next to the QR Code. It will start with 86512'\n"
    "3. 'Where is my IMEI?' expandable help section is visible\n"
    "4. Field accepts up to 15 digits and shows count progress (e.g. 6/15)\n"
    "5. Tapping 'Continue' with valid IMEI advances to Connect Tracker screen"
),
"test_data": (
    "imei: 860211234567890\n"
    "format: 15 digits, starts with 86512\n"
    "account: testuser@example.com"
),
"postconditions": "- IMEI is validated and user advances to Connect Tracker screen",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-009",
"title": "Verify displaying Device Not Found screen when a manually entered IMEI is not registered to the account",
"preconditions": (
    "- User is on the 'Enter IMEI' screen\n"
    "- IMEI to be entered is valid 15-digit format but NOT registered to test account"
),
"steps": (
    "1. Enter a valid-format 15-digit IMEI not registered to the account\n"
    "2. Tap 'Continue'\n"
    "3. Observe the result screen"
),
"expected": (
    "1. 'Device Not Found' screen appears\n"
    "2. Message: 'The IMEI provided is not registered to your account. Please check the code or contact support.'\n"
    "3. 'Try Again', 'Contact Support', and 'Exit Setup' buttons are visible"
),
"test_data": (
    "imei: 860219999999999\n"
    "status: Valid format, not registered to account"
),
"postconditions": "- No device is linked to the account",
"status": "Not Run", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-010",
"title": "Verify Connect Tracker screen displays GPS optimisation guidance and network connectivity status after valid IMEI",
"preconditions": (
    "- User has scanned or entered a valid registered IMEI\n"
    "- Tracker is powered and connected"
),
"steps": (
    "1. Observe the Connect Tracker screen after successful IMEI validation\n"
    "2. Observe the GPS optimisation note\n"
    "3. Observe the Connectivity section\n"
    "4. Tap 'Continue'"
),
"expected": (
    "1. Screen title 'Connect Tracker' with PHASE 01 / 05 label\n"
    "2. Tracker device image is shown\n"
    "3. GPS Optimisation note: 'Park in an open area for initial satellite synchronisation. Avoid underground'\n"
    "4. Connectivity section shows NETWORK: CONNECTED with green indicator\n"
    "5. Buttons visible: 'Continue', 'Scan QR Code', 'Save & Exit'\n"
    "5. Tapping 'Continue' advances to Phase 2"
),
"test_data": "tracker_imei: 860211234567890\nnetwork_status: Connected",
"postconditions": "- User advances to Phase 2: Rapid Photo Capture",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-011",
"title": "Verify displaying Save and Exit confirmation modal when user taps Save and Exit during setup",
"preconditions": (
    "- User is in any active phase of the Vehicle Setup flow\n"
    "- User has not yet completed setup"
),
"steps": (
    "1. Tap 'Save & Exit' button on any setup screen\n"
    "2. Observe the modal that appears\n"
    "3. Tap 'Continue Vehicle Setup'\n"
    "4. Tap 'Save & Exit' again and tap 'Exit and save for later'"
),
"expected": (
    "1. Modal appears with title 'Exit Setup?'\n"
    "2. Message: 'Your progress will be saved. You can pick up from where you left off next time you return.'\n"
    "3. Two buttons visible: 'Continue Vehicle Setup' and 'Exit and save for later'\n"
    "3. Tapping 'Continue Vehicle Setup' dismisses modal and returns user to the same setup screen\n"
    "4. Tapping 'Exit and save for later' exits setup and saves progress"
),
"test_data": "app_account: testuser@example.com\nsetup_progress: Any phase in progress",
"postconditions": (
    "- When 'Exit and save for later' is tapped: setup progress is saved to the account\n"
    "- When 'Continue Vehicle Setup' is tapped: user returns to the current setup screen"
),
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Design update 24/04/2026: Save & Exit now shows confirmation modal to prevent accidental exit"
},

{
"tc_id": "DVS-012",
"title": "Verify resuming Vehicle Setup from saved progress when user returns after exiting and saving",
"preconditions": (
    "- User has previously started Vehicle Setup and used 'Exit and save for later'\n"
    "- Setup progress was saved at a specific phase\n"
    "- User logs back into the app"
),
"steps": (
    "1. Log into the Telemax app with the same account\n"
    "2. Navigate to or tap the in-progress setup entry\n"
    "3. Observe which screen the setup resumes at"
),
"expected": (
    "1. App shows the incomplete vehicle setup entry\n"
    "2. Tapping the entry resumes setup from the exact phase where the user exited\n"
    "3. Previously captured data (e.g. IMEI, photos) is preserved\n"
    "3. User does not need to restart from Phase 1"
),
"test_data": "app_account: testuser@example.com\nsaved_phase: Phase 2 (example)",
"postconditions": "- Setup continues from the saved checkpoint with all prior data intact",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

# ── PHASE 2: Rapid Photo Capture ─────────────────────────────────
{
"tc_id": "DVS-013",
"title": "Verify capturing odometer reading successfully with a clear and well-lit dashboard image",
"preconditions": (
    "- User has completed Phase 1 (Device Connection)\n"
    "- User is on the Odometer Reading capture screen\n"
    "- Camera permission is granted\n"
    "- Vehicle dashboard is visible and odometer display is readable\n"
    "- 'SYNCING IN BACKGROUND' banner is active at top of screen"
),
"steps": (
    "1. Observe the Odometer Reading screen\n"
    "2. Position phone to frame the odometer display within the viewfinder\n"
    "3. Tap the camera button at the bottom of the screen\n"
    "4. Observe the processing state\n"
    "5. Observe the result"
),
"expected": (
    "1. Screen shows PHASE 02 / 05 label and 'Odometer Reading' title\n"
    "2. Instruction: 'Position your phone to clearly capture the current mileage on the dashboard display'\n"
    "3. PRO TIP displayed: 'If conditions are too bright, try shading the plate with your body to avoid direct reflection'\n"
    "4. Camera button triggers image capture; 'PROCESSING IMAGE' overlay appears\n"
    "5. After processing, app advances to the next capture screen (VIN)"
),
"test_data": "odometer_value: Current reading on test vehicle dashboard\nlight_condition: Normal indoor/outdoor lighting",
"postconditions": "- Odometer reading is captured and stored; user advances to VIN capture screen",
"status": "PENDING", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Blocked by CQ-002 — no Figma screen defined for complete OCR failure (unreadable image). Expected result step 5 assumes success; failure path unknown.",
"pending": True
},

{
"tc_id": "DVS-014",
"title": "Verify displaying processing image overlay state while odometer photo is being analysed",
"preconditions": (
    "- User has tapped the camera button on the Odometer Reading screen\n"
    "- Image has been captured and is being processed"
),
"steps": (
    "1. Tap the camera capture button\n"
    "2. Observe the screen immediately after capture"
),
"expected": (
    "1. 'PROCESSING IMAGE' overlay banner appears over the camera viewfinder\n"
    "2. Camera viewfinder still shows the last captured frame\n"
    "3. User cannot tap the camera button again while processing\n"
    "4. Overlay disappears once processing completes"
),
"test_data": "capture_trigger: Camera button tap",
"postconditions": "- Processing state resolves to either success (advance) or retry prompt",
"status": "Not Run", "priority": "Medium", "type": "UI",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-015",
"title": "Verify capturing VIN plate successfully with a clear and unobstructed image",
"preconditions": (
    "- User has completed Odometer capture\n"
    "- User is on the VIN capture screen\n"
    "- Vehicle VIN plate is visible and unobscured\n"
    "- Camera permission is granted"
),
"steps": (
    "1. Observe the VIN capture screen\n"
    "2. Position phone to frame the VIN plate within the viewfinder\n"
    "3. Tap the camera button"
),
"expected": (
    "1. Screen shows PHASE 02 / 05 label and 'VIN' title\n"
    "2. Instruction: 'Position the VIN plate within the frame. Ensure it\\'s clearly visible without glare.'\n"
    "3. PRO TIP: 'If conditions are too bright, try shading the plate with your body to avoid direct reflection'\n"
    "4. Camera captures and processes the VIN plate image\n"
    "5. App advances to Registration Plate capture screen"
),
"test_data": "vin_plate: Visible, unobscured VIN plate on test vehicle",
"postconditions": "- VIN image captured with high confidence; user advances to Registration Plate screen",
"status": "PENDING", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Blocked by CQ-002 — no Figma screen defined for complete OCR failure on VIN. Manual entry fallback not confirmed.",
"pending": True
},

{
"tc_id": "DVS-016",
"title": "Verify capturing registration plate successfully with a clear and unobstructed image",
"preconditions": (
    "- User has completed VIN capture\n"
    "- User is on the Registration Plate capture screen\n"
    "- Vehicle registration plate is visible and readable\n"
    "- Camera permission is granted"
),
"steps": (
    "1. Observe the Registration Plate screen\n"
    "2. Position phone to frame the registration plate within the viewfinder\n"
    "3. Tap the camera button"
),
"expected": (
    "1. Screen shows PHASE 02 / 05 label and 'Registration Plate' title\n"
    "2. Instruction: 'Position the vehicle registration plate within the frame. Ensure it\\'s clearly visible without glare.'\n"
    "3. PRO TIP is displayed\n"
    "4. Camera captures and processes the registration plate image\n"
    "5. App advances to Phase 3: Start Engine & Sync"
),
"test_data": "rego_plate: 123ABC\nplate_condition: Clean, unobscured",
"postconditions": "- Registration plate captured; all Phase 2 photos complete; user advances to Phase 3",
"status": "PENDING", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Blocked by CQ-002 — no Figma screen defined for complete OCR failure on registration plate. Manual entry fallback not confirmed.",
"pending": True
},

{
"tc_id": "DVS-017",
"title": "Verify syncing in background banner is displayed throughout all Phase 2 photo capture screens",
"preconditions": (
    "- User has completed Phase 1 and is in Phase 2\n"
    "- Background sync is running"
),
"steps": (
    "1. Navigate to Odometer Reading screen\n"
    "2. Observe the top banner\n"
    "3. Navigate to VIN screen\n"
    "4. Observe the top banner\n"
    "5. Navigate to Registration Plate screen\n"
    "6. Observe the top banner"
),
"expected": (
    "1. 'SYNCING IN BACKGROUND' banner with sync icon is visible at the very top of the screen\n"
    "2. Banner persists across all three Phase 2 capture screens\n"
    "3. Banner does not block any interactive UI elements"
),
"test_data": "phase: 02 — Rapid Photo Capture",
"postconditions": "- Banner remains visible throughout Phase 2",
"status": "Not Run", "priority": "Medium", "type": "UI",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

# ── PHASE 3: Start Engine & Sync ─────────────────────────────────
{
"tc_id": "DVS-018",
"title": "Verify displaying Start Engine preconditions screen with handbrake and park mode checklist before sync",
"preconditions": (
    "- User has completed Phase 2 (Rapid Photo Capture)\n"
    "- User is on the Start Engine screen (Phase 3)"
),
"steps": (
    "1. Observe the Start Engine screen\n"
    "2. Observe the precondition checklist items\n"
    "3. Tap 'I\\'ve Turned On The Engine'"
),
"expected": (
    "1. Screen shows PHASE 03 / 05 label and 'Start Engine' title\n"
    "2. Subtitle: 'Ensure the vehicle is secured before proceeding with the diagnostic sync'\n"
    "3. Checklist shows two items with green checkmarks:\n"
    "   - 'Handbrake ON — VEHICLE MUST BE STATIONARY'\n"
    "   - 'Park Mode — TRANSMISSION IN \\\"P\\\" OR NEUTRAL'\n"
    "4. 'I\\'ve Turned On The Engine' button is visible and tappable\n"
    "4. 'Save & Exit' link is visible\n"
    "4. Tapping the button starts the sync countdown"
),
"test_data": "vehicle_state: Handbrake on, transmission in Park",
"postconditions": "- Sync countdown begins; user sees Syncing Vehicle Data screen",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-019",
"title": "Verify syncing vehicle data successfully and displaying all OBD data within the 60 second countdown",
"preconditions": (
    "- User has tapped 'I\\'ve Turned On The Engine'\n"
    "- Engine is running\n"
    "- OBD device is correctly plugged into the vehicle's OBD-II port\n"
    "- Vehicle is stationary with handbrake on"
),
"steps": (
    "1. Observe the Syncing Vehicle Data screen\n"
    "2. Observe the countdown timer and progress indicators\n"
    "3. Wait for sync to complete within 60 seconds\n"
    "4. Observe the Sync Successful screen"
),
"expected": (
    "1. Screen shows countdown timer (e.g. '45s REMAINING')\n"
    "2. Title: 'Syncing Vehicle Data' — 'Establishing secure handshake between hardware and Telemax Connect'\n"
    "3. SYNC progress bar visible (e.g. 65%)\n"
    "4. Live data appears as it syncs: VIN (with cross-referencing status), ODOMETER (km from OBD), BATTERY VOLTAGE\n"
    "4. 'Waiting For Data...' shows for items not yet received\n"
    "5. Sync Successful screen shows '100% COMPLETE' with all data confirmed:\n"
    "   - VIN IDENTIFIED with full VIN string\n"
    "   - ODOMETER (KM) value\n"
    "   - BATTERY VOLTAGE with 'Optimal' label\n"
    "5. 'Continue' and 'Save & Exit' buttons are visible"
),
"test_data": (
    "vehicle_vin: 1HGCM82633A004352\n"
    "odometer: 42891 km (read from OBD)\n"
    "battery_voltage: 14.2V"
),
"postconditions": (
    "- Sync data is stored to the vehicle record\n"
    "- User advances to Phase 4 when 'Continue' is tapped"
),
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-020",
"title": "Verify displaying Sync Successful screen with all three verified data points after successful OBD sync",
"preconditions": (
    "- Sync has completed successfully within 60 seconds\n"
    "- All three data points received: VIN, Odometer, Battery Voltage"
),
"steps": (
    "1. Observe the Sync Successful screen\n"
    "2. Verify all displayed values\n"
    "3. Tap 'Continue'"
),
"expected": (
    "1. Progress indicator shows '100% COMPLETE'\n"
    "2. Title: 'Sync Successful'\n"
    "3. Subtitle: 'Vehicle connection is stable and all encrypted data clusters have been verified'\n"
    "4. VIN IDENTIFIED field shows full VIN with blue checkmark\n"
    "5. ODOMETER (KM) shows value with blue checkmark\n"
    "6. BATTERY VOLTAGE shows value with 'Optimal' label and blue checkmark\n"
    "7. Tapping 'Continue' advances to Phase 4: Vehicle Review"
),
"test_data": (
    "vehicle_vin: 1HGCM82633A004352\n"
    "odometer_km: 42891.4\n"
    "battery_voltage: 14.2V / Optimal"
),
"postconditions": "- User advances to Phase 4 (Vehicle Review)",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-021",
"title": "Verify displaying Sync Timeout screen when no OBD data is received within 60 seconds",
"preconditions": (
    "- User is on Syncing Vehicle Data screen\n"
    "- OBD device is not transmitting data (e.g. engine not fully running, device not properly seated)\n"
    "- 60-second timer is running"
),
"steps": (
    "1. Wait for the 60-second countdown to expire without OBD data received\n"
    "2. Observe the resulting screen\n"
    "3. Tap 'Retry'\n"
    "4. On second timeout, tap 'Skip'"
),
"expected": (
    "1. After 60 seconds with no OBD data: 'Sync Timeout' screen appears\n"
    "2. Message: 'We haven\\'t received OBD data from your vehicle in over 60 seconds. You can still continue with the next steps.'\n"
    "3. Options available: Retry (resets 60-second timer), Skip (continues without OBD data, no verified badge)\n"
    "3. Tapping 'Retry' resets the 60s timer and returns to Syncing screen\n"
    "4. Tapping 'Skip' advances to Phase 4 without OBD-verified data"
),
"test_data": "obd_state: No data transmitted (OBD device not seated or engine not running)\ntimeout: 60 seconds",
"postconditions": "- Skip: user advances to Phase 4 without VIN/odometer verified badge",
"status": "Not Run", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

# ── PHASE 4: Vehicle Review ───────────────────────────────────────
{
"tc_id": "DVS-022",
"title": "Verify displaying Review Details screen with all vehicle data correctly populated from photos and OBD sync",
"preconditions": (
    "- User has completed Phases 1–3 successfully\n"
    "- All photos captured with high OCR confidence\n"
    "- OBD sync was successful"
),
"steps": (
    "1. Observe the Review Details screen after completing Phase 3\n"
    "2. Verify each field is populated"
),
"expected": (
    "1. Screen shows PHASE 04 / 05 label and 'Review Details' title\n"
    "2. NAME field shows vehicle name (e.g. '344ITU Ford Ranger')\n"
    "3. VEHICLE field shows make/model/year (e.g. 'Ford Ranger 2024')\n"
    "4. FUEL TYPE & CAPACITY shows detected values (e.g. 'Diesel 80L')\n"
    "5. TRANSMISSION shows detected value (e.g. 'Automatic')\n"
    "6. VEHICLE IDENTIFICATION NUMBER (VIN) shows detected VIN with green verified icon\n"
    "7. REGISTRATION NUMBER shows plate value (e.g. '123ABC')\n"
    "8. CURRENT ODOMETER shows OBD-read value (e.g. '15,420 km')\n"
    "9. 'Confirm Vehicle Details', 'Edit Details', and 'Save & Exit' are visible"
),
"test_data": (
    "vehicle_name: 344ITU Ford Ranger\n"
    "make_model_year: Ford Ranger 2024\n"
    "fuel: Diesel 80L\n"
    "transmission: Automatic\n"
    "vin: GEDG41515494G15SR\n"
    "rego: 123ABC\n"
    "odometer: 15420 km"
),
"postconditions": "- Review Details displays all populated vehicle data for user confirmation",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-023",
"title": "Verify editing vehicle details successfully and returning to Review Details screen with updated values",
"preconditions": (
    "- User is on the Review Details screen\n"
    "- At least one field needs correction (e.g. incorrect fuel type)"
),
"steps": (
    "1. Tap 'Edit Details' on the Review Details screen\n"
    "2. Observe the Edit Vehicle screen\n"
    "3. Change 'Fuel type' from 'Petrol' to 'Diesel' using the dropdown\n"
    "4. Tap 'Confirm Vehicle Details'\n"
    "5. Observe the Review Details screen"
),
"expected": (
    "1. Edit Vehicle screen loads with all current values pre-filled\n"
    "2. Fields shown: Vehicle Name, Make, Model, Build Year, Fuel type, Fuel Capacity (L), Transmission\n"
    "3. All fields use dropdown selectors\n"
    "4. Changing fuel type to 'Diesel' updates the dropdown selection\n"
    "5. Tapping 'Confirm Vehicle Details' saves changes and returns to Review Details\n"
    "5. Review Details now shows 'Diesel' in the FUEL TYPE & CAPACITY field"
),
"test_data": (
    "original_fuel: Petrol\n"
    "updated_fuel: Diesel\n"
    "vehicle_name: 344ITU Ford Ranger"
),
"postconditions": "- Updated fuel type is reflected on Review Details screen",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-024",
"title": "Verify tapping Cancel on Edit Vehicle screen returns user to Review Details without saving changes",
"preconditions": (
    "- User is on the Edit Vehicle screen having made changes to one or more fields"
),
"steps": (
    "1. On the Edit Vehicle screen, change a field value (e.g. change Build Year)\n"
    "2. Tap 'Cancel'\n"
    "3. Observe the Review Details screen"
),
"expected": (
    "1. Tapping 'Cancel' returns user to the Review Details screen\n"
    "2. The changed field value is NOT saved — Review Details shows the original value\n"
    "3. No data is lost or overwritten"
),
"test_data": "changed_field: Build Year\noriginal_value: 2024\nchanged_value: 2023",
"postconditions": "- Review Details shows original values unchanged",
"status": "Not Run", "priority": "Medium", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-025",
"title": "Verify displaying VIN field error with retake button when OCR confidence is low due to partially obscured image",
"preconditions": (
    "- User is on the Review Details screen\n"
    "- VIN photo was captured with low OCR confidence (e.g. partially obscured, blurry)"
),
"steps": (
    "1. Observe the Review Details screen after a low-confidence VIN photo\n"
    "2. Observe the VIN field state\n"
    "3. Tap 'Retake Photo'"
),
"expected": (
    "1. VEHICLE IDENTIFICATION NUMBER (VIN) field shows in red text: 'ERROR: IMAGE PARTIALLY OBSCURED'\n"
    "2. A red 'RETAKE PHOTO' button appears below the VIN field\n"
    "3. 'Confirm Vehicle Details' button is disabled or shown in grey (cannot confirm with unresolved error)\n"
    "4. Tapping 'RETAKE PHOTO' opens the camera for VIN capture only\n"
    "4. Other fields (rego, odometer, etc.) retain their captured values"
),
"test_data": "vin_image: Partially obscured VIN plate photo\nocr_confidence: Low",
"postconditions": "- Camera opens for VIN retake only; other captured data is preserved",
"status": "PENDING", "priority": "High", "type": "Negative",
"environment": "Mobile (iOS 16+ / Android 12+)",
"notes": "Blocked by CQ-003 — Figma does not clarify whether 'Confirm Vehicle Details' is fully disabled (greyed) or shows a validation error on tap when VIN error is unresolved.",
"pending": True
},

{
"tc_id": "DVS-026",
"title": "Verify retaking a single photo field does not overwrite or lose data captured in other fields",
"preconditions": (
    "- User is on Review Details with a low-confidence VIN requiring retake\n"
    "- Other fields (rego, odometer) have been successfully captured"
),
"steps": (
    "1. Tap 'RETAKE PHOTO' for the VIN field\n"
    "2. Capture a new clear VIN image\n"
    "3. Return to Review Details screen\n"
    "4. Verify all other fields"
),
"expected": (
    "1. Camera opens for VIN capture only\n"
    "2. New VIN image is processed and replaces only the VIN field value\n"
    "3. After successful retake, VIN field shows the correct value with green tick\n"
    "4. REGISTRATION NUMBER, ODOMETER, FUEL TYPE, and all other fields remain unchanged\n"
    "4. 'Confirm Vehicle Details' button becomes active"
),
"test_data": "retake_field: VIN\nother_fields: rego=123ABC, odometer=15420km (must be preserved)",
"postconditions": "- VIN is updated; all other fields unchanged; user can confirm vehicle details",
"status": "Not Run", "priority": "Medium", "type": "Edge Case",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-027",
"title": "Verify confirming vehicle details successfully from Review Details screen and advancing to Phase 5",
"preconditions": (
    "- User is on Review Details screen\n"
    "- All fields are populated with no unresolved errors\n"
    "- VIN shows green verified tick"
),
"steps": (
    "1. Review all details on the Review Details screen\n"
    "2. Tap 'Confirm Vehicle Details'"
),
"expected": (
    "1. All fields display correct data with no red error states\n"
    "2. Tapping 'Confirm Vehicle Details' saves the vehicle record\n"
    "3. App advances to Phase 5: Final Health Check"
),
"test_data": (
    "vehicle_name: 344ITU Ford Ranger\n"
    "vin: GEDG41515494G15SR\n"
    "rego: 123ABC\n"
    "odometer: 15420 km"
),
"postconditions": (
    "- Vehicle record is saved to the account with all confirmed details\n"
    "- User advances to Phase 5"
),
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

# ── PHASE 5: Final Health Check ───────────────────────────────────
{
"tc_id": "DVS-028",
"title": "Verify displaying Final Health Check screen with all three hardware module status indicators passing",
"preconditions": (
    "- User has completed Phases 1–4\n"
    "- Tracker is active and broadcasting\n"
    "- Mobile has LTE connection\n"
    "- GPS fix has been established\n"
    "- Vehicle power supply is stable"
),
"steps": (
    "1. Observe the Final Health Check screen\n"
    "2. Observe each module status row\n"
    "3. Tap 'Back'\n"
    "4. Return and tap 'Complete Setup'"
),
"expected": (
    "1. Screen shows PHASE 05 / 05 label and 'Final Health Check' title\n"
    "2. Description: 'System diagnostics confirm all hardware modules are broadcasting within optimal parameters'\n"
    "3. Mobile Network row: OK badge, 'LTE Connection Active', signal strength bar chart\n"
    "4. GPS Position row: OK badge, '3D Fix Established', map thumbnail showing location\n"
    "5. Power Supply row: OK badge, '13.8V Nominal Input'\n"
    "6. 'Complete Setup' and 'Back' buttons are visible\n"
    "6. Tapping 'Back' returns user to Phase 4 Review Details\n"
    "6. Tapping 'Complete Setup' advances to Ready to Drive screen"
),
"test_data": (
    "mobile_network: LTE active\n"
    "gps_fix: 3D Fix established\n"
    "power_supply: 13.8V"
),
"postconditions": "- All health checks pass; user advances to Ready to Drive screen",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-029",
"title": "Verify displaying Ready to Drive screen with vehicle summary after completing all 5 phases",
"preconditions": (
    "- User has tapped 'Complete Setup' on the Final Health Check screen\n"
    "- All 5 phases completed successfully"
),
"steps": (
    "1. Tap 'Complete Setup' on Final Health Check\n"
    "2. Observe the Ready to Drive screen\n"
    "3. Verify all displayed values"
),
"expected": (
    "1. Blue circular checkmark icon is displayed prominently\n"
    "2. Title: 'Ready to Drive'\n"
    "3. Label: 'PHASE 5 OF 5 COMPLETE'\n"
    "4. VEHICLE NAME shows the confirmed vehicle name (e.g. '344ITU Ford Ranger')\n"
    "5. TRACKER STATUS: 'Fully Synced' with three-dot menu\n"
    "6. HEALTH SCAN: '100%'\n"
    "7. NEXT SYNC: 'Real-time'\n"
    "8. Buttons visible: 'Go To Vehicles' and 'Set Up Another Vehicle' link"
),
"test_data": "vehicle_name: 344ITU Ford Ranger\nall_phases: Completed",
"postconditions": (
    "- Vehicle is fully set up and visible in the fleet\n"
    "- Vehicle status: Active / Fully Synced"
),
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-030",
"title": "Verify navigating to vehicle list successfully when tapping Go To Vehicles from Ready to Drive screen",
"preconditions": (
    "- User is on the Ready to Drive screen\n"
    "- Setup has been completed for at least one vehicle"
),
"steps": (
    "1. Tap 'Go To Vehicles' button\n"
    "2. Observe the destination screen"
),
"expected": (
    "1. App navigates to the Vehicles list screen\n"
    "2. The newly set-up vehicle appears in the list with status 'Fully Synced'\n"
    "3. Vehicle details (name, tracker status) match the values entered during setup"
),
"test_data": "vehicle_name: 344ITU Ford Ranger",
"postconditions": "- User is on the Vehicles list; new vehicle is visible and active",
"status": "Not Run", "priority": "High", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

{
"tc_id": "DVS-031",
"title": "Verify tapping Set Up Another Vehicle from Ready to Drive screen restarts the Vehicle Setup flow",
"preconditions": (
    "- User is on the Ready to Drive screen\n"
    "- User needs to set up a second vehicle"
),
"steps": (
    "1. Tap 'Set Up Another Vehicle' link\n"
    "2. Observe the destination screen"
),
"expected": (
    "1. App navigates back to the Welcome to Vehicle Onboarding screen\n"
    "2. Welcome screen is fresh — no pre-filled data from the previous setup\n"
    "3. All 5 phases are shown as not started"
),
"test_data": "prior_vehicle: 344ITU Ford Ranger (already set up)",
"postconditions": "- New setup session starts; previously set up vehicle is unaffected",
"status": "Not Run", "priority": "Medium", "type": "Functional",
"environment": "Mobile (iOS 16+ / Android 12+)", "notes": ""
},

]

# ════════════════════════════════════════════════════════════════
#  PHASE SEPARATOR MARKERS
# ════════════════════════════════════════════════════════════════
PHASE_LABELS = {
    "DVS-001": "PHASE 01 / 05 — Device Connection",
    "DVS-013": "PHASE 02 / 05 — Rapid Photo Capture",
    "DVS-018": "PHASE 03 / 05 — Start Engine & Sync",
    "DVS-022": "PHASE 04 / 05 — Vehicle Review",
    "DVS-028": "PHASE 05 / 05 — Final Health Check",
}

# ════════════════════════════════════════════════════════════════
#  CLARIFY REQUIREMENTS DATA
# ════════════════════════════════════════════════════════════════
CLARIFY_REQUIREMENTS = [
    {
        "ref":        "CQ-001",
        "tc_id":      "DVS-003",
        "source":     "Figma: Phase 1 / Connectivity Check — Connection Required screen",
        "conflict":   "The 'Continue Anyway' button is visible on the Connection Required screen but the Figma does not define what behaviour is restricted or degraded when the user proceeds without an internet connection. It is unclear whether photo capture is queued for later upload or fails entirely.",
        "question":   "When the user taps 'Continue Anyway' without internet, which specific features are restricted during setup? Can photos still be captured and queued for upload when connection is restored, or does the lack of internet block photo processing entirely?",
        "priority":   "High",
        "raised_by":  "",
        "raised_date":"2026-04-25",
        "answer":     "",
        "status":     "Open",
        "resolved":   "",
    },
    {
        "ref":        "CQ-002",
        "tc_id":      "DVS-013, DVS-015, DVS-016",
        "source":     "Figma: Phase 2 / Odometer Reading, VIN, Registration Plate screens",
        "conflict":   "The Figma defines a 'low confidence' OCR state (warning icon + Retake button shown in Phase 4 Review Details) but does NOT define what happens when OCR fails completely — i.e. image is too dark, blurry, or unreadable and no value can be extracted at all. There is no manual text entry fallback visible in any Phase 2 screen.",
        "question":   "If OCR cannot extract any value from a captured photo (complete failure, not low confidence), what is the expected behaviour? Is there a manual text entry fallback for odometer, VIN, and registration plate? Or does the app require the user to retake until a value is extracted?",
        "priority":   "High",
        "raised_by":  "",
        "raised_date":"2026-04-25",
        "answer":     "",
        "status":     "Open",
        "resolved":   "",
    },
    {
        "ref":        "CQ-003",
        "tc_id":      "DVS-025",
        "source":     "Figma: Phase 4 / Review Details — VIN error state (ERROR: IMAGE PARTIALLY OBSCURED)",
        "conflict":   "The Figma shows the 'Confirm Vehicle Details' button on the Review Details screen when a VIN error is present, but it is not clear whether the button is fully disabled (greyed out and non-tappable) or remains enabled and shows a validation error/toast when tapped.",
        "question":   "When the VIN field shows 'ERROR: IMAGE PARTIALLY OBSCURED', is the 'Confirm Vehicle Details' button: (a) fully disabled and greyed out, OR (b) still tappable but shows a validation error message on tap?",
        "priority":   "High",
        "raised_by":  "",
        "raised_date":"2026-04-25",
        "answer":     "",
        "status":     "Open",
        "resolved":   "",
    },
    {
        "ref":        "CQ-004",
        "tc_id":      "DVS-006",
        "source":     "Figma: Phase 1 / Scan QR Code — design update note 24/04/2026",
        "conflict":   "The design update note states that scanning a 'serial number (different format)' QR code keeps the user on the Scan QR Code page with an error banner. It is unclear whether this 'stays on page' behaviour applies ONLY to the serial number QR format on the same tracker device, or to ALL unrecognised / non-IMEI QR codes (e.g. a QR from a completely unrelated product).",
        "question":   "Does the 'stays on page with error banner' behaviour apply only to the tracker's serial number QR code, or to any QR code that does not match the 15-digit IMEI format starting with 86512? What happens if a completely unrelated QR code (e.g. from packaging, URL) is scanned?",
        "priority":   "Medium",
        "raised_by":  "",
        "raised_date":"2026-04-25",
        "answer":     "",
        "status":     "Open",
        "resolved":   "",
    },
    {
        "ref":        "CQ-005",
        "tc_id":      "DVS-021",
        "source":     "Figma: Phase 3 / Sync Timeout screen annotation",
        "conflict":   "The Sync Timeout annotation mentions 'Skip continues without OBD data. Relies on OBD-II data already captured during Steps 4-8.' It is unclear what 'Steps 4-8' refers to in the user flow and whether any partial OBD data captured before the timeout is retained and displayed in Phase 4 Review Details.",
        "question":   "When the user chooses Skip on Sync Timeout, is any partial OBD data (e.g. VIN detected but odometer not received) retained and shown in Phase 4 Review Details? Or is the review populated only from photo OCR data in this case?",
        "priority":   "Medium",
        "raised_by":  "",
        "raised_date":"2026-04-25",
        "answer":     "",
        "status":     "Open",
        "resolved":   "",
    },
]

# ════════════════════════════════════════════════════════════════
#  BUILD CQ SHEET
# ════════════════════════════════════════════════════════════════
def build_cq_sheet(wb):
    wc = wb.create_sheet("Clarify Requirements")

    cq_col_widths = {1:10, 2:20, 3:32, 4:46, 5:46, 6:10, 7:14, 8:14, 9:30, 10:10, 11:14}
    for col, width in cq_col_widths.items():
        wc.column_dimensions[get_column_letter(col)].width = width

    # Title
    wc.merge_cells("A1:K1")
    wc["A1"] = "Clarify Requirements — Device & Vehicle Setup Flow"
    wc["A1"].font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
    wc["A1"].fill      = fill(NAVY)
    wc["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wc.row_dimensions[1].height = 28

    # Legend
    wc.merge_cells("A2:K2")
    wc["A2"] = (
        "ORANGE ROW = TC blocked — do not execute until Status = Answered or Closed  |  "
        "SOURCE FORMAT: Figma: 'Phase / Screen'  |  "
        "PRIORITY: High = blocks TC  |  Medium = partial block  |  Low = low-risk assumption"
    )
    wc["A2"].font      = Font(name="Calibri", size=8, italic=True, color=GREY_TEXT)
    wc["A2"].fill      = fill("F2F2F2")
    wc["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wc.row_dimensions[2].height = 22

    # Column headers
    cq_headers = [
        "Ref #", "Related TC ID", "Source / Reference",
        "Conflict Description", "Question to Client",
        "Priority", "Raised By", "Raised Date",
        "Answer / Resolution", "Status", "Resolved Date"
    ]
    for c, h in enumerate(cq_headers, 1):
        cell = wc.cell(row=3, column=c, value=h)
        cell.font      = hdr_font(size=10)
        cell.fill      = fill(BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    wc.row_dimensions[3].height = 22
    wc.freeze_panes = "A4"
    wc.auto_filter.ref = f"A3:{get_column_letter(11)}3"

    # CQ rows
    for i, cq in enumerate(CLARIFY_REQUIREMENTS, 4):
        is_high = cq["priority"] == "High"
        row_fill = fill(ORANGE) if is_high else fill(BLUE_LIGHT)
        values = [
            cq["ref"], cq["tc_id"], cq["source"],
            cq["conflict"], cq["question"],
            cq["priority"], cq["raised_by"], cq["raised_date"],
            cq["answer"], cq["status"], cq["resolved"],
        ]
        for c, val in enumerate(values, 1):
            cell = wc.cell(row=i, column=c, value=val)
            cell.border    = thin_border()
            cell.alignment = wrap_align()
            cell.fill      = row_fill
            cell.font      = body_font(bold=(c == 1))
        wc.row_dimensions[i].height = 80

    # Status flow note
    last = len(CLARIFY_REQUIREMENTS) + 5
    wc.merge_cells(start_row=last, start_column=1, end_row=last, end_column=11)
    note = "STATUS FLOW:  Open  →  Answered  →  Closed   |   Once Answered: update the linked TC, remove orange highlight from TC row, then set Status to Closed."
    cell = wc.cell(row=last, column=1, value=note)
    cell.font      = Font(name="Calibri", size=8, italic=True, color=GREY_TEXT)
    cell.fill      = fill("F2F2F2")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wc.row_dimensions[last].height = 14


# ════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Device & Vehicle Setup"

# ── Column widths ──
col_widths = {1:12, 2:40, 3:36, 4:36, 5:36, 6:28, 7:32, 8:12, 9:10, 10:13, 11:22, 12:32}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Row 1: main title ──
ws.merge_cells("A1:L1")
ws["A1"] = "Test Cases: Device & Vehicle Setup Flow"
ws["A1"].font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
ws["A1"].fill      = fill(NAVY)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Row 2: metadata ──
ws.merge_cells("A2:L2")
ws["A2"] = "Module: DVS  |  Environment: Mobile (iOS 16+ / Android 12+)  |  Version: 2.0  |  Status: In Progress"
ws["A2"].font      = Font(name="Calibri", size=9, color=GREY_TEXT)
ws["A2"].fill      = fill("F2F2F2")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Row 3: column headers ──
headers = ["TC ID","Title","Preconditions","Steps (Action)","Expected Result",
           "Test Data","Postconditions","Status","Priority","Type","Environment","Notes"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font      = hdr_font(size=10)
    cell.fill      = fill(BLUE_MID)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[3].height = 22

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:{get_column_letter(12)}3"

# ── Write rows ──
row = 4
for tc in TESTCASES:
    if tc["tc_id"] in PHASE_LABELS:
        phase_sep(ws, row, PHASE_LABELS[tc["tc_id"]])
        row += 1
    write_tc(ws, row, tc)
    row += 1

# ── Summary table ──
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.cell(row=row, column=1, value="SUMMARY").font = Font(name="Calibri", size=11, bold=True, color=WHITE)
ws.cell(row=row, column=1).fill      = fill(NAVY)
ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

sum_headers = ["TC ID", "Title", "Phase", "Type", "Priority", "Status"]
for c, h in enumerate(sum_headers, 1):
    cell = ws.cell(row=row, column=c, value=h)
    cell.font      = hdr_font()
    cell.fill      = fill(BLUE_MID)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[row].height = 18
row += 1

phase_map = {
    "DVS-001": "Phase 1","DVS-002": "Phase 1","DVS-003": "Phase 1",
    "DVS-004": "Phase 1","DVS-005": "Phase 1","DVS-006": "Phase 1",
    "DVS-007": "Phase 1","DVS-008": "Phase 1","DVS-009": "Phase 1",
    "DVS-010": "Phase 1","DVS-011": "Phase 1","DVS-012": "Phase 1",
    "DVS-013": "Phase 2","DVS-014": "Phase 2","DVS-015": "Phase 2",
    "DVS-016": "Phase 2","DVS-017": "Phase 2",
    "DVS-018": "Phase 3","DVS-019": "Phase 3","DVS-020": "Phase 3","DVS-021": "Phase 3",
    "DVS-022": "Phase 4","DVS-023": "Phase 4","DVS-024": "Phase 4",
    "DVS-025": "Phase 4","DVS-026": "Phase 4","DVS-027": "Phase 4",
    "DVS-028": "Phase 5","DVS-029": "Phase 5","DVS-030": "Phase 5","DVS-031": "Phase 5",
}

for i, tc in enumerate(TESTCASES):
    row_fill = fill(ORANGE) if tc.get("pending") else (fill(BLUE_LIGHT) if i % 2 == 0 else fill(WHITE))
    data = [
        tc["tc_id"],
        tc["title"][:80] + ("..." if len(tc["title"]) > 80 else ""),
        phase_map.get(tc["tc_id"], ""),
        tc["type"],
        tc["priority"],
        tc["status"],
    ]
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = body_font(bold=(c == 1))
        cell.fill      = row_fill
        cell.border    = thin_border()
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

# ── Legend row ──
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
legend = (
    "LEGEND: ORANGE ROW = TC has open Clarify Requirement (CQ) — do not execute until resolved  |  "
    "See 'Clarify Requirements' sheet for open questions  |  "
    "Status flow: Not Run → In Progress → Pass / Fail / Blocked / Skip / Pending"
)
cell = ws.cell(row=row, column=1, value=legend)
cell.font      = Font(name="Calibri", size=8, italic=True, color=GREY_TEXT)
cell.fill      = fill("F2F2F2")
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 14

# ── Build CQ sheet ──
build_cq_sheet(wb)

# ── Save ──
out = r"d:\Manual Testing\Manual Test Cases Rules\Test Case for Device and vehicle setup flow\TestCases_DeviceVehicleSetup_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total TCs: {len(TESTCASES)}")
print(f"Total CQs: {len(CLARIFY_REQUIREMENTS)}")
pending = [tc['tc_id'] for tc in TESTCASES if tc.get('pending')]
print(f"PENDING TCs ({len(pending)}): {', '.join(pending)}")
