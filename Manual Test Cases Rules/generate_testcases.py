import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Device & Vehicle Setup"

# ---------- column headers ----------
headers = [
    "TC ID", "Title", "Preconditions", "Steps (Action)",
    "Expected Result", "Test Data", "Postconditions",
    "Status", "Priority", "Type", "Environment", "Notes"
]

# ---------- styles ----------
header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", fgColor="1F4E79")
phase_font    = Font(name="Calibri", bold=True, size=11)
phase_fill    = PatternFill("solid", fgColor="D6E4F0")
cell_font     = Font(name="Calibri", size=10)
wrap_align    = Alignment(wrap_text=True, vertical="top")
center_align  = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------- column widths ----------
col_widths = [12, 55, 45, 45, 45, 35, 40, 12, 12, 14, 28, 30]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- header row ----------
ws.row_dimensions[1].height = 30
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# ---------- test case data ----------
# Each entry: (TC_ID, Title, Preconditions, Steps, Expected, TestData, Post, Status, Priority, Type, Env, Notes)
NL = "\n"  # line break inside cell

test_cases = [
    # ── PHASE 1 separator ──
    ("PHASE 1 — Device Connection", "", "", "", "", "", "", "", "", "", "", ""),

    ("DVS-001",
     "Verify navigating to Connectivity Check when tapping Get Started on Welcome screen",
     "1. User is logged in to the Telemax app\n2. Vehicle Setup has not been previously started\n3. Device: iOS 16+ or Android 12+\n4. App is on latest version",
     "1. Open the Telemax app\n2. Navigate to Vehicle Setup\n3. View Welcome to Telemax Vehicle Onboarding screen\n4. Tap \"Get Started\"",
     "1. Telemax app opens and home screen is displayed\n2. Welcome screen loads with 5 phases listed (Device Connection, Photo Capture, Start Engine & Sync, Vehicle Review, Health Check)\n3. \"Get Started\" and \"Cancel\" buttons are visible\n4. App navigates to Connectivity Check screen (Phase 01/05)",
     "account: testuser@telemax.com\npassword: TestPass123!",
     "User is on the Connectivity Check screen\nSetup progress is initialized",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-002",
     "Verify exiting Vehicle Setup when tapping Cancel on Welcome screen",
     "1. User is logged in to the Telemax app\n2. User is on the Welcome to Telemax Vehicle Onboarding screen",
     "1. View Welcome screen with 5-phase overview\n2. Tap \"Cancel\"",
     "1. Welcome screen displays with Cancel link visible below Get Started\n2. User is navigated away from Vehicle Setup back to the previous screen",
     "account: testuser@telemax.com\npassword: TestPass123!",
     "No setup progress is saved\nUser returns to previous screen",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-003",
     "Verify proceeding to QR Code instructions when mobile connectivity shows Status: Connected",
     "1. User is on Connectivity Check screen (Phase 01/05)\n2. Mobile device has active internet connection\n3. Signal strength is sufficient",
     "1. View Connectivity Check screen\n2. Confirm status shows \"Status: Connected\" with green indicator\n3. Tap \"Start Setup\"",
     "1. Screen displays \"Status: Connected\" with green checkmark icon, latency value, and signal bars\n2. \"Start Setup\" button is enabled\n3. App navigates to How to Scan QR Code screen",
     "network: WiFi or 4G/5G active\naccount: testuser@telemax.com",
     "User is on How to Scan QR Code screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-004",
     "Verify displaying Connection Required error when mobile device has no internet on Connectivity Check",
     "1. User is on Connectivity Check screen (Phase 01/05)\n2. Mobile device has no active internet connection (airplane mode or no signal)",
     "1. Disable internet on the test device (airplane mode)\n2. Navigate to Connectivity Check screen\n3. View the connectivity status",
     "1. Connectivity Check screen loads\n2. Status shows \"Connection Required\" with red/warning indicator\n3. Error message states internet connection is needed for photo processing and vehicle data\n4. Three options shown: \"Wait For Signal\", \"Continue Anyway\", \"Exit Setup\"",
     "network: disabled (airplane mode)\naccount: testuser@telemax.com",
     "User remains on Connectivity Check screen with Connection Required state",
     "Not Run", "High", "Negative", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-005",
     "Verify waiting on Connectivity Check screen when tapping Wait For Signal with no connection",
     "1. User is on Connectivity Check screen\n2. Connection Required error is displayed\n3. Device has no internet connection",
     "1. View Connection Required state on Connectivity Check screen\n2. Tap \"Wait For Signal\"\n3. Re-enable internet on device",
     "1. Connection Required error and three options are visible\n2. App remains on Connectivity Check screen; no navigation occurs\n3. Status updates to \"Status: Connected\" once signal is detected",
     "network: initially disabled, then re-enabled\naccount: testuser@telemax.com",
     "User remains on Connectivity Check screen until connection is restored",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-006",
     "Verify continuing setup when tapping Continue Anyway despite Connection Required on Connectivity Check",
     "1. User is on Connectivity Check screen\n2. Connection Required error is displayed\n3. Device has no internet connection",
     "1. View Connection Required error on Connectivity Check screen\n2. Tap \"Continue Anyway\"",
     "1. Connection Required state displays with Continue Anyway option visible\n2. App navigates forward to How to Scan QR Code screen despite no connection",
     "network: disabled (airplane mode)\naccount: testuser@telemax.com",
     "User proceeds to next step without internet; photo upload features may be limited",
     "Not Run", "Medium", "Negative", "Mobile (iOS 16+ / Android 12+)", "Offline-dependent steps may fail later in the flow"),

    ("DVS-007",
     "Verify navigating to QR Code scanner when tapping Start Scan on How to Scan QR Code screen",
     "1. User is on How to Scan QR Code screen\n2. Camera permission is granted on device",
     "1. View How to Scan QR Code screen with scanning instructions\n2. Read instructions: QR code is next to IMEI number, 15-digit number starting with 8602\n3. Tap \"Start Scan\"",
     "1. Instructions shown with QR code placement diagram and IMEI format description\n2. \"Watch Installation Video\", \"Start Scan\", \"Save & Exit\", \"Can't Scan? Enter IMEI Manually\" options are visible\n3. Camera view opens on Scan QR Code screen with scanning frame",
     "camera_permission: granted\naccount: testuser@telemax.com",
     "User is on Scan QR Code screen with camera active",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-008",
     "Verify connecting tracker successfully when valid QR code is scanned on Scan QR Code screen",
     "1. User is on Scan QR Code screen\n2. Camera is active with scanning frame visible\n3. OBD device with valid QR code is available\n4. IMEI on device is registered to the test account",
     "1. View Scan QR Code screen with camera active\n2. Position the tracker's QR code within the scanning frame\n3. Wait for automatic scan detection",
     "1. Camera view shows scanning frame with tip to place finger over bottom QR code to prevent scanning both at the same time\n2. QR code is detected automatically\n3. App navigates to Connect Tracker screen showing successful connection with GPS Optimization tips and Connectivity status",
     "device_imei: 860211234567890\nqr_code: valid QR on OBD device\naccount: testuser@telemax.com",
     "Tracker is connected and recognized\nUser is on Connect Tracker screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-009",
     "Verify displaying incorrect QR code error banner when non-tracker QR code is scanned",
     "1. User is on Scan QR Code screen\n2. Camera is active\n3. A QR code that is not a valid tracker IMEI is available",
     "1. View Scan QR Code screen with active camera\n2. Position an invalid/non-tracker QR code within the scanning frame\n3. Wait for scan detection",
     "1. Camera scanning frame is active and ready\n2. Invalid QR code is detected\n3. Error banner appears: \"Incorrect QR Code Scanned — The IMEI number is a 15-digit number starting with 8602\"\n4. User remains on Scan QR Code screen to retry",
     "qr_code: invalid (e.g. product QR, URL QR)\naccount: testuser@telemax.com",
     "User remains on Scan QR Code screen\nNo tracker is connected",
     "Not Run", "High", "Negative", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-010",
     "Verify navigating to Enter IMEI screen when tapping Can't Scan Enter IMEI Manually",
     "1. User is on Scan QR Code screen or How to Scan QR Code screen\n2. User cannot scan QR code (damaged label or poor camera conditions)",
     "1. View Scan QR Code screen\n2. Tap \"Can't Scan? Enter IMEI Manually\" link",
     "1. \"Can't Scan? Enter IMEI Manually\" link is visible below the scanning frame\n2. App navigates to Enter IMEI screen with IMEI input field and helper text (\"Where is my IMEI?\")",
     "account: testuser@telemax.com",
     "User is on Enter IMEI screen",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-011",
     "Verify connecting tracker successfully when valid registered IMEI is entered manually",
     "1. User is on Enter IMEI screen\n2. IMEI number of registered OBD device is known\n3. Device is registered to the test account",
     "1. View Enter IMEI screen with input field\n2. Type the valid 15-digit IMEI number (starting with 8602)\n3. Tap \"Continue\"",
     "1. IMEI input field is visible with format hint and \"Where is my IMEI?\" helper\n2. 15-digit IMEI is entered into the field\n3. App navigates to Connect Tracker screen showing successful connection",
     "imei: 860211234567890\naccount: testuser@telemax.com",
     "Tracker is connected and recognized\nUser is on Connect Tracker screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-012",
     "Verify displaying Device Not Found error when an unregistered IMEI is entered manually",
     "1. User is on Enter IMEI screen\n2. An IMEI that is not registered to the test account is available",
     "1. View Enter IMEI screen\n2. Enter a valid-format but unregistered 15-digit IMEI\n3. Tap \"Continue\"",
     "1. IMEI field accepts the 15-digit input\n2. App submits the IMEI\n3. Device Not Found screen appears: \"The IMEI provided is not registered to your account. Please check the code or contact support.\"\n4. Options shown: \"Try Again\", \"Contact Support\", \"Exit Setup\"",
     "imei: 860299999999999\naccount: testuser@telemax.com",
     "No tracker connected\nUser is on Device Not Found screen",
     "Not Run", "High", "Negative", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-013",
     "Verify returning to Enter IMEI screen when tapping Try Again on Device Not Found",
     "1. User is on Device Not Found screen\n2. Previous IMEI entry was unregistered",
     "1. View Device Not Found screen with Try Again, Contact Support, Exit Setup options\n2. Tap \"Try Again\"",
     "1. Device Not Found screen displays with error message and all three options\n2. App navigates back to Enter IMEI screen\n3. IMEI input field is cleared and ready for new entry",
     "account: testuser@telemax.com",
     "User is back on Enter IMEI screen with empty field",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-014",
     "Verify saving setup progress and resuming from saved point when tapping Save & Exit during Phase 1",
     "1. User is mid-flow in Phase 1 (any screen after Welcome)\n2. User has a valid session",
     "1. Navigate to any Phase 1 screen (e.g. How to Scan QR Code)\n2. Tap \"Save & Exit\"\n3. Re-open Vehicle Setup",
     "1. Save & Exit option is visible on the current screen\n2. Progress is saved and user exits to previous screen\n3. Vehicle Setup resumes from the exact screen where the user exited",
     "account: testuser@telemax.com",
     "Setup progress persisted in app state\nUser can re-enter and continue from saved point",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", "Per Figma: next session starts from where user exited"),

    # ── PHASE 2 separator ──
    ("PHASE 2 — Rapid Photo Capture", "", "", "", "", "", "", "", "", "", "", ""),

    ("DVS-015",
     "Verify capturing odometer reading photo successfully when image is clear and undistorted",
     "1. Phase 1 (Device Connection) is completed\n2. User is on Odometer Reading screen (Phase 02/05)\n3. Syncing in Background banner is visible\n4. Camera permission is granted\n5. Vehicle dashboard is accessible and odometer is visible",
     "1. View Odometer Reading screen with camera frame and Pro Tip\n2. Position phone to align odometer display within the camera frame\n3. Tap the camera capture button",
     "1. Screen shows Phase 02/05, camera frame with overlay, and Pro Tip: \"If conditions are too bright, try shading the plate with your body to avoid direct reflection\"\n2. Camera frame aligns with odometer display\n3. Processing Image overlay appears on captured image; app proceeds to VIN capture screen",
     "odometer_reading: clearly visible on dashboard\nlighting: normal indoor/outdoor",
     "Odometer image captured and sent for OCR processing\nUser proceeds to VIN capture screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", "Syncing in Background banner remains throughout Phase 2"),

    ("DVS-016",
     "Verify displaying Processing Image overlay when odometer photo is being processed after capture",
     "1. User is on Odometer Reading screen\n2. Photo has just been captured",
     "1. Tap the camera capture button on Odometer Reading screen\n2. Observe the screen immediately after capture",
     "1. Camera button is tapped\n2. \"Processing Image\" overlay badge appears on the captured image thumbnail\n3. Processing indicator remains until OCR result is returned",
     "N/A — state triggered by camera capture",
     "Image processing is in progress in background",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-017",
     "Verify capturing VIN plate photo successfully when VIN is clearly visible without glare",
     "1. Odometer photo has been captured\n2. User is on VIN screen (Phase 02/05)\n3. Vehicle VIN plate is accessible\n4. Camera permission is granted",
     "1. View VIN screen with camera frame and Pro Tip\n2. Position phone so VIN plate is within the frame and clearly visible\n3. Tap the camera capture button",
     "1. VIN screen shows Phase 02/05 label, camera frame, Pro Tip about avoiding glare\n2. VIN plate is positioned correctly in frame\n3. Photo is captured and Processing Image overlay appears; app proceeds to Registration Plate screen",
     "vin_plate: clearly visible, no glare\nlighting: adequate",
     "VIN image captured and sent for OCR processing\nUser proceeds to Registration Plate screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-018",
     "Verify capturing registration plate photo successfully when plate is clearly visible without glare",
     "1. VIN photo has been captured\n2. User is on Registration Plate screen (Phase 02/05)\n3. Vehicle registration plate is accessible\n4. Camera permission is granted",
     "1. View Registration Plate screen with camera frame and Pro Tip\n2. Position phone so registration plate is within the frame and clearly visible\n3. Tap the camera capture button",
     "1. Registration Plate screen shows Phase 02/05 label, camera frame, Pro Tip about avoiding glare\n2. Registration plate is positioned correctly in frame\n3. Photo is captured and Processing Image overlay appears; Phase 2 photo capture is complete",
     "registration_plate: clearly visible, no glare\nlighting: adequate",
     "Registration plate image captured and sent for OCR\nAll Phase 2 photos are captured; user proceeds to Phase 3",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    # ── PHASE 3 separator ──
    ("PHASE 3 — Start Engine & Sync", "", "", "", "", "", "", "", "", "", "", ""),

    ("DVS-019",
     "Verify initiating vehicle data sync when engine is on and I've Turned On The Engine is tapped",
     "1. Phase 2 photo capture is completed\n2. User is on Start Engine screen (Phase 03/05)\n3. Vehicle is stationary with handbrake ON\n4. Transmission is in P or Neutral\n5. OBD device is connected to vehicle",
     "1. View Start Engine screen showing Handbrake ON and Park Mode checklist\n2. Turn on the vehicle engine\n3. Confirm Handbrake ON and Park Mode checkboxes show green ticks\n4. Tap \"I've Turned On The Engine\"",
     "1. Start Engine screen shows two prerequisites: Handbrake ON and Park Mode both with green checkmarks\n2. Engine is running\n3. Both prerequisites confirmed with green ticks\n4. App navigates to Syncing Vehicle Data screen with 60-second countdown timer",
     "vehicle_state: engine on, handbrake on, transmission in P\nobd_connected: true",
     "Sync process has started\nUser is on Syncing Vehicle Data screen with countdown active",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", "Vehicle must be physically present for this test"),

    ("DVS-020",
     "Verify displaying Sync Successful with verified vehicle data when OBD sync completes within 60 seconds",
     "1. User is on Syncing Vehicle Data screen\n2. 60-second countdown is active\n3. OBD device is connected and transmitting data\n4. Vehicle engine is on",
     "1. View Syncing Vehicle Data screen with countdown timer and sync progress\n2. Observe real-time data: VIN cross-referencing, odometer reading, battery voltage\n3. Wait for sync to reach 100%",
     "1. Screen shows circular countdown (e.g. 45s remaining), SYNC progress bar, VIN cross-referencing, odometer km, battery voltage, and Waiting For Data status\n2. All data fields populate progressively\n3. App navigates to Sync Successful screen showing 100% complete, VIN Identified, Odometer (KM), Battery Voltage with Optimal status, and Continue button",
     "obd_data: VIN=1HGCM82633A004\nodometer: 42891 km\nbattery: 14.2V",
     "OBD data synced and verified\nVIN badge is marked as Verified\nUser is on Sync Successful screen",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-021",
     "Verify displaying Sync Timeout screen and resetting timer when Retry is tapped after 60 seconds without OBD data",
     "1. User is on Syncing Vehicle Data screen\n2. OBD device is not transmitting data (engine off or device fault)\n3. 60-second countdown has expired with no data received",
     "1. View Syncing Vehicle Data screen with countdown active\n2. Allow 60-second timer to expire without OBD data\n3. View Sync Timeout screen\n4. Tap \"Retry\"",
     "1. Syncing screen shows countdown timer counting down\n2. Timer reaches 0 with no OBD data received\n3. Sync Timeout screen appears: \"We haven't received OBD data from your vehicle in over 60 seconds\" with Retry, Skip & Continue, Save & Exit options\n4. App returns to Syncing Vehicle Data screen with a fresh 60-second countdown",
     "obd_data: none (device not transmitting)",
     "Sync retried with fresh timer\nUser is back on Syncing Vehicle Data screen",
     "Not Run", "High", "Edge Case", "Mobile (iOS 16+ / Android 12+)", "Retry resets 60-second timer only; does not reset previously captured photo data"),

    ("DVS-022",
     "Verify proceeding to Phase 4 without OBD verification badge when Skip & Continue is tapped on Sync Timeout",
     "1. User is on Sync Timeout screen\n2. 60-second timer has expired with no OBD data received\n3. Phase 2 photo OCR data is available",
     "1. View Sync Timeout screen with Retry, Skip & Continue, Save & Exit options\n2. Tap \"Skip & Continue\"",
     "1. Sync Timeout screen is visible with all three options\n2. App proceeds to next phase (Vehicle Review) using only photo OCR data from Phase 2\n3. VIN is NOT marked as Verified (no verified badge shown on Review Details)",
     "obd_data: none\nocr_data: available from Phase 2 photos",
     "Setup continues without OBD data\nVerified badge is not present on VIN field\nUser is on Vehicle Review screen",
     "Not Run", "High", "Negative", "Mobile (iOS 16+ / Android 12+)", "Per Figma: no Verified badge when skipping OBD sync"),

    # ── PHASE 4 separator ──
    ("PHASE 4 — Vehicle Review", "", "", "", "", "", "", "", "", "", "", ""),

    ("DVS-023",
     "Verify confirming vehicle details and proceeding when all fields show high confidence on Review Details",
     "1. Phase 3 sync is completed successfully\n2. User is on Review Details screen (Phase 04/05)\n3. All OCR fields returned with high confidence (green ticks shown)",
     "1. View Review Details screen with all vehicle fields populated\n2. Verify all fields show green tick (high confidence): Name, Vehicle, Fuel Type & Capacity, Transmission, VIN, Registration Number, Odometer\n3. Tap \"Confirm Vehicle Details\"",
     "1. Review Details screen shows all fields: Name, Vehicle, Fuel Type & Capacity, Transmission, VIN with green checkmark, Registration Number, Current Odometer\n2. All fields display green ticks indicating high confidence OCR results\n3. App proceeds to Phase 5 (Health Check)",
     "vehicle_name: 344ITU Ford Ranger\nmake: Ford\nmodel: Ranger 2024\nfuel_type: Diesel\nfuel_capacity: 80L\ntransmission: Automatic\nvin: GEDG41515494G15SR\nregistration: 123ABC\nodometer: 15420 km",
     "Vehicle details confirmed and saved\nUser proceeds to Phase 5",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-024",
     "Verify navigating to Edit Vehicle screen when tapping Edit Details on Review Details",
     "1. User is on Review Details screen (Phase 04/05)\n2. Vehicle details are populated from OCR/OBD",
     "1. View Review Details screen with all populated vehicle fields\n2. Tap \"Edit Details\" link",
     "1. Review Details screen is fully visible with Confirm Vehicle Details, Edit Details, and Save & Exit options\n2. App navigates to Edit Vehicle screen showing editable fields: Vehicle Name, Make, Model, Build Year, Fuel Type, Fuel Capacity, Transmission — all as dropdowns",
     "account: testuser@telemax.com",
     "User is on Edit Vehicle screen with pre-filled values from Review Details",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-025",
     "Verify returning to Review Details with updated values when Confirm Vehicle Details is tapped on Edit Vehicle",
     "1. User is on Edit Vehicle screen\n2. One or more fields have been changed (e.g. Fuel Type from Petrol to Diesel)",
     "1. View Edit Vehicle screen with pre-filled dropdown fields\n2. Change Fuel Type dropdown from Petrol to Diesel\n3. Tap \"Confirm Vehicle Details\"",
     "1. Edit Vehicle screen shows all editable fields (Vehicle Name, Make, Model, Build Year, Fuel Type, Fuel Capacity, Transmission)\n2. Fuel Type dropdown updates to Diesel\n3. App returns to Review Details screen showing the updated Fuel Type value (Diesel)",
     "fuel_type_before: Petrol\nfuel_type_after: Diesel",
     "Updated vehicle details reflected on Review Details screen\nNo other fields were changed",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-026",
     "Verify returning to Review Details without changes when Cancel is tapped on Edit Vehicle",
     "1. User is on Edit Vehicle screen\n2. User has or has not made changes to fields",
     "1. View Edit Vehicle screen with all dropdown fields\n2. Optionally change one field value\n3. Tap \"Cancel\"",
     "1. Edit Vehicle screen is visible with Confirm Vehicle Details and Cancel options\n2. Field value may have been changed on screen\n3. App returns to Review Details with original values unchanged — no edits are saved",
     "account: testuser@telemax.com",
     "Review Details retains original values\nUser is back on Review Details screen",
     "Not Run", "Medium", "Functional", "Mobile (iOS 16+ / Android 12+)", ""),

    ("DVS-027",
     "Verify displaying low confidence warning and Retake Photo button when VIN image has partial obstruction on Review Details",
     "1. User is on Review Details screen (Phase 04/05)\n2. VIN photo was captured with partial obstruction (low OCR confidence)\n3. OCR returned a low-confidence result for the VIN field",
     "1. View Review Details screen\n2. Observe the VIN field result from OCR\n3. Identify the low-confidence warning on the VIN field",
     "1. Review Details screen loads with all vehicle fields\n2. VIN field shows \"ERROR: IMAGE PARTIALLY OBSCURED\" in red text with a warning indicator\n3. A red \"Retake Photo\" button appears next to the VIN field\n4. Other high-confidence fields retain their green ticks unchanged",
     "vin_photo: partially obscured image\nocr_confidence: low",
     "VIN field flagged as low confidence\nOther fields unaffected",
     "Not Run", "High", "Negative", "Mobile (iOS 16+ / Android 12+)", "Per Figma: low-confidence fields show warning icon + Retake Photo; high-confidence fields show green tick"),

    ("DVS-028",
     "Verify retaking VIN photo without losing other field data when Retake Photo is tapped on low-confidence VIN field",
     "1. User is on Review Details screen\n2. VIN field shows \"ERROR: IMAGE PARTIALLY OBSCURED\" with Retake Photo button\n3. All other fields have valid OCR data",
     "1. View Review Details screen with VIN low-confidence error and red Retake Photo button\n2. Tap \"Retake Photo\" on the VIN field\n3. Capture a clear VIN photo\n4. Return to Review Details",
     "1. Retake Photo button is visible on the VIN row only\n2. Camera opens for VIN photo only; other screens are not revisited\n3. New VIN photo is captured successfully\n4. Review Details updates VIN field with new OCR result and shows green tick; all other field values remain unchanged",
     "vin_photo_retake: clear, fully visible VIN plate\nexisting_ocr_data: odometer=15420km, registration=123ABC (must remain intact)",
     "VIN field updated with high-confidence value\nAll other field values preserved",
     "Not Run", "High", "Functional", "Mobile (iOS 16+ / Android 12+)", "Per Figma: retaking one photo must not lose data from other fields"),
]

# ---------- open clarification map ----------
# TC IDs that have at least one Open CQ, mapped to the CQ ref(s) blocking them
open_cq_map = {
    "DVS-006":  "Blocked by CQ-002",
    "DVS-009":  "Blocked by CQ-001",
    "DVS-012":  "Blocked by CQ-003",
    "DVS-021":  "Blocked by CQ-004",
    "DVS-022":  "Blocked by CQ-005",
    "DVS-027":  "Blocked by CQ-006",
    "DVS-028":  "Blocked by CQ-006",
}

pending_fill   = PatternFill("solid", fgColor="FFC000")   # orange — open question
pending_font   = Font(name="Calibri", size=10, bold=True, color="000000")

# ---------- write rows ----------
row_num = 2
for tc in test_cases:
    is_phase  = tc[0].startswith("PHASE")
    tc_id     = tc[0]
    is_pending = tc_id in open_cq_map

    ws.row_dimensions[row_num].height = 15 if is_phase else 120

    # build the row as a mutable list so we can patch Status and Notes
    row_values = list(tc)
    if is_pending:
        row_values[7]  = "PENDING"                          # col H — Status
        cq_note        = open_cq_map[tc_id]
        existing_note  = row_values[11].strip()
        row_values[11] = (cq_note + "\n" + existing_note).strip()  # col L — Notes

    for col, value in enumerate(row_values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = border

        if is_phase:
            cell.fill      = PatternFill("solid", fgColor="2E75B6")
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif is_pending:
            cell.fill = pending_fill
            if col in (8, 9, 10):   # Status, Priority, Type
                cell.font      = pending_font
                cell.alignment = center_align
            else:
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = wrap_align
        else:
            cell.font = cell_font
            if col in (8, 9, 10):
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    row_num += 1

# ---------- freeze panes & auto-filter ----------
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:L1"

# ══════════════════════════════════════════════════════════
#  SHEET 2 — Clarify Requirements
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Clarify Requirements")

cq_headers = [
    "Ref #", "Related TC ID", "Source / Reference",
    "Conflict Description", "Question to Client",
    "Priority", "Raised By", "Raised Date",
    "Answer / Resolution", "Status", "Resolved Date"
]

cq_widths = [10, 16, 24, 45, 45, 12, 14, 14, 45, 12, 14]

for i, w in enumerate(cq_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# --- header row ---
ws2.row_dimensions[1].height = 30
for col, h in enumerate(cq_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

# --- legend rows (rows 2-4: status definitions) ---
legend_fill   = PatternFill("solid", fgColor="FFF2CC")
legend_font   = Font(name="Calibri", size=9.5, italic=True, color="595959")
legend_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)

legend_lines = [
    "STATUS GUIDE →   Open: waiting for answer  |  Answered: reply received, update TC + Postman request  |  Closed: TC updated and verified",
    "SOURCE FORMAT →   User Story: 'Story US-042'  |  Figma: 'Phase 1 / Scan QR Code'  |  API: 'POST /api/v1/...'  ||  PRIORITY: High = blocks TC writing | Medium = partially writable | Low = low-risk  ||  API CQ rows included here only when client requests combined export.",
]

for r_idx, line in enumerate(legend_lines, 2):
    ws2.row_dimensions[r_idx].height = 18
    for col in range(1, len(cq_headers) + 1):
        cell = ws2.cell(row=r_idx, column=col)
        cell.fill   = legend_fill
        cell.border = legend_border
        if col == 1:
            cell.value     = line
            cell.font      = legend_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

# merge legend cells across all columns (rows 2 and 3 only)
for r_idx in range(2, 4):
    ws2.merge_cells(start_row=r_idx, start_column=1,
                    end_row=r_idx,   end_column=len(cq_headers))

# --- sub-header row (row 4) — repeat column labels, acts as filter anchor ---
ws2.row_dimensions[4].height = 20
for col in range(1, len(cq_headers) + 1):
    cell = ws2.cell(row=4, column=col, value=cq_headers[col - 1])
    cell.fill      = PatternFill("solid", fgColor="D6E4F0")
    cell.font      = Font(name="Calibri", size=9.5, bold=True, color="1F4E79")
    cell.border    = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws2.auto_filter.ref = "A4:K4"

# --- sample open clarification rows (start at row 5) ---
open_fill   = PatternFill("solid", fgColor="FFE6E6")   # red tint  — Open
ans_fill    = PatternFill("solid", fgColor="E2EFDA")   # green tint — Answered
closed_fill = PatternFill("solid", fgColor="EDEDED")   # grey       — Closed

cq_data = [
    # ── UI clarifications ────────────────────────────────────────────────────
    # Ref#, TC ID, Phase/Screen · Endpoint, Conflict, Question, Priority, By, Date, Answer, Status, Resolved
    ("CQ-001", "DVS-009",
     "Figma: Phase 1 / Scan QR Code",
     "The error banner text for an incorrect QR scan is not defined in Figma. The banner appears but no copy (message text) is specified.",
     "What is the exact error message text shown when an incorrect QR code is scanned?",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-002", "DVS-006",
     "Figma: Phase 1 / Connectivity Check",
     "Figma shows 'Continue Anyway' as an option when Connection Required, but does not define which features are restricted or degraded when proceeding without internet.",
     "Which specific features or steps are restricted/unavailable when the user continues without an internet connection?",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-003", "DVS-012",
     "Figma: Phase 1 / Enter IMEI",
     "The Device Not Found screen shows 'Contact Support' but no support destination (URL, email, or in-app chat) is defined in Figma.",
     "Where does 'Contact Support' navigate to — an in-app chat, email, or external URL?",
     "Medium", "", "2026-04-25", "", "Open", ""),

    ("CQ-004", "DVS-021",
     "Figma: Phase 3 / Sync Timeout",
     "The Figma note states 'Skip & Continue proceeds without OBD data' but it is unclear whether the user receives any warning or confirmation before skipping.",
     "Does tapping 'Skip & Continue' show a confirmation dialog before proceeding, or does it navigate immediately?",
     "Medium", "", "2026-04-25", "", "Open", ""),

    ("CQ-005", "DVS-022",
     "Figma: Phase 3 / Sync Timeout",
     "The Figma note states 'No Verified badge' when skipping OBD sync, but does not specify whether the missing badge is communicated to the user on the Review Details screen.",
     "Is the absence of the Verified badge explained to the user on the Review Details screen, or is it silently omitted?",
     "Medium", "", "2026-04-25", "", "Open", ""),

    ("CQ-006", "DVS-027, DVS-028",
     "Figma: Phase 4 / Review Details",
     "Low-confidence flagging shows 'ERROR: IMAGE PARTIALLY OBSCURED' for VIN, but it is not defined whether the same error type applies to Odometer and Registration Plate fields.",
     "Does the low-confidence / partial obstruction error apply to all photo fields (Odometer, VIN, Registration Plate), or only VIN?",
     "High", "", "2026-04-25", "", "Open", ""),

    # ── API clarifications ───────────────────────────────────────────────────
    ("CQ-007", "API-002",
     "API: POST /api/v1/device/validate-imei",
     "API spec defines 200 (success) and 404 (not found) but does not document the response body schema for the 404 error case. Field names and message format are unknown.",
     "What are the exact field names and values in the 404 error response body when an unregistered IMEI is submitted? (e.g. error code, message string)",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-008", "API-003",
     "API: POST /api/v1/device/validate-imei",
     "API spec does not explicitly state whether the validate-imei endpoint requires authentication. It is unclear if an unauthenticated call returns 401 or proceeds publicly.",
     "Does POST /api/v1/device/validate-imei require a Bearer token? If yes, what is the exact 401 response body format?",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-009", "API-006, API-007, API-008",
     "API: POST /api/v1/vehicle/photos/upload",
     "The spec does not define the maximum accepted file size for photo uploads, nor the accepted MIME types. It is unclear what status code and message is returned when a file exceeds the limit.",
     "What is the maximum file size allowed for photo uploads? What MIME types are accepted? What status code and message is returned when the limit is exceeded?",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-010", "API-009",
     "API: POST /api/v1/vehicle/sync",
     "The spec does not define the side effects of starting an OBD sync. It is unclear whether calling this endpoint multiple times creates duplicate sync jobs or replaces the existing active job.",
     "Is POST /api/v1/vehicle/sync idempotent? If called twice with the same device_id, does it create a second sync job or return the existing job_id?",
     "Medium", "", "2026-04-25", "", "Open", ""),

    ("CQ-011", "API-010, API-011",
     "API: GET /api/v1/vehicle/sync/{job_id}",
     "The spec does not define the polling interval or maximum number of polls before the client should stop. The timeout status code (408) is assumed but not confirmed in the spec.",
     "What is the recommended polling interval for sync status? What is the exact HTTP status code and response body returned when the 60-second timeout is reached?",
     "High", "", "2026-04-25", "", "Open", ""),

    ("CQ-012", "API-013",
     "API: POST /api/v1/vehicle/confirm",
     "The spec does not document side effects of confirming vehicle details. It is unclear whether this endpoint triggers an email notification, a webhook, or any background processing.",
     "Does POST /api/v1/vehicle/confirm trigger any side effects (email, webhook, background job)? If yes, what are they and how should they be verified in testing?",
     "Medium", "", "2026-04-25", "", "Open", ""),

    ("CQ-013", "API-014",
     "API: PUT /api/v1/vehicle/{vehicle_id}",
     "The spec does not state whether all fields must be sent in the PUT body or only the fields being changed (partial update). It is ambiguous whether this is a full replace or a partial patch.",
     "Does PUT /api/v1/vehicle/{vehicle_id} require the full vehicle object in the body, or does it accept partial updates with only the changed fields?",
     "High", "", "2026-04-25", "", "Open", ""),
]

status_fills = {"Open": open_fill, "Answered": ans_fill, "Closed": closed_fill}

for r_idx, row_data in enumerate(cq_data, 5):  # data starts at row 5
    ws2.row_dimensions[r_idx].height = 80
    row_fill = status_fills.get(row_data[9], PatternFill())

    for col, value in enumerate(row_data, 1):
        cell = ws2.cell(row=r_idx, column=col, value=value)
        cell.border = border
        cell.fill   = row_fill

        # status and priority centered
        if col in (6, 10):
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            if col == 10:
                status_color = {"Open": "C00000", "Answered": "375623", "Closed": "595959"}
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color=status_color.get(value, "000000"))
            else:
                cell.font = Font(name="Calibri", size=10)
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font      = Font(name="Calibri", size=10)

# --- freeze panes below sub-header ---
ws2.freeze_panes = "A5"

# ---------- save ----------
output_path = r"d:\Manual Testing\Manual Test Cases Rules\TestCases_DeviceVehicleSetup_v4.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
